from __future__ import annotations

import calendar
import hashlib
import json
from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

from packages.common.storage import connect_sqlite


SHANGHAI_TZ = ZoneInfo("Asia/Shanghai")
CONTRIBUTION_TYPES = {"regular", "night", "ppp"}
CONTRIBUTION_SCORE_CAP = 5.0
CONTRIBUTION_HIGH_VIEW_BONUS = 0.5
PUBLISHER_KINDS = {"human", "human_unmapped", "odaily_ai", "other_ai", "pending_ai"}
SHIFT_DEFINITIONS = {
    "three": (
        ("morning", "早班", time(8, 0), time(13, 30), 0),
        ("middle", "中班", time(13, 30), time(19, 30), 0),
        ("late", "晚班", time(19, 30), time(1, 0), 1),
    ),
    "two": (
        ("morning", "早班", time(8, 0), time(16, 0), 0),
        ("late", "晚班", time(16, 0), time(0, 0), 1),
    ),
}
SOURCE_LABELS = {
    "odaily": "Odaily",
    "blockbeats": "BlockBeats",
    "panews": "PANews",
    "jinse": "金色财经",
}
QUALITY_FIRST_WEEK = date(2026, 8, 3)
QUALITY_THRESHOLD_MULTIPLIER = 1.5
QUALITY_KPI_PER_ITEM = 0.2
QUALITY_REGULAR_SOURCE_ACCOUNTS = (
    "LinChen91162689", "ki_young_ju", "BTC__options", "EricBalchunas", "EleanorTerrett",
    "zhusu", "evilcos", "sunyuchentron", "NateGeraci", "ag_dwf", "intotheblock", "weremeow",
    "VitalikButerin", "santimentfeed", "0xENAS", "cz_binance", "JSeyff", "elonmusk",
    "CryptoHayes", "zachxbt", "Rewkang", "jessepollak", "PrimordialAA", "Matrixport_CN",
    "10x_Research", "im23pds", "brian_armstrong", "ali_charts", "haydenzadams",
)
QUALITY_EXTERNAL_MEDIA_URLS = (
    "https://www.coindesk.com/",
    "https://cointelegraph.com/",
    "https://www.theblock.co/",
    "https://decrypt.co/",
)
QUALITY_EXCLUSION_GROUPS = (
    {"key": "btc_liquidation", "terms": ("BTC", "爆仓")},
    {"key": "eth_liquidation", "terms": ("ETH", "爆仓")},
    {"key": "sol_liquidation", "terms": ("SOL", "爆仓")},
)
QUALITY_OVERRIDE_VALUES = {"none", "include", "exclude"}


SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS newsflash_roster (
    person_key text PRIMARY KEY,
    display_name text NOT NULL,
    duty_enabled integer NOT NULL DEFAULT 0,
    contributor_enabled integer NOT NULL DEFAULT 0,
    active integer NOT NULL DEFAULT 1,
    created_at text NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at text NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS newsflash_roster_aliases (
    alias_normalized text PRIMARY KEY,
    alias_display text NOT NULL,
    person_key text NOT NULL REFERENCES newsflash_roster(person_key) ON DELETE CASCADE,
    created_at text NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS newsflash_operation_facts (
    source_item_id text PRIMARY KEY,
    operator_raw text,
    publisher_kind text,
    publisher_person_key text REFERENCES newsflash_roster(person_key) ON DELETE SET NULL,
    publisher_locked integer NOT NULL DEFAULT 0,
    view_count integer,
    is_pushed integer,
    pushed_at text,
    is_contribution integer NOT NULL DEFAULT 0,
    contributor_person_key text REFERENCES newsflash_roster(person_key) ON DELETE SET NULL,
    contribution_type text NOT NULL DEFAULT 'regular',
    quality_override text NOT NULL DEFAULT 'none',
    source_snapshot_at text,
    attribution_checked_at text,
    created_at text NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at text NOT NULL DEFAULT CURRENT_TIMESTAMP,
    CHECK (contribution_type IN ('regular','night','ppp')),
    CHECK (publisher_kind IS NULL OR publisher_kind IN ('human','human_unmapped','odaily_ai','other_ai','pending_ai'))
);
CREATE INDEX IF NOT EXISTS idx_newsflash_operation_operator
ON newsflash_operation_facts(publisher_person_key, publisher_kind, updated_at DESC);
CREATE INDEX IF NOT EXISTS idx_newsflash_operation_contribution
ON newsflash_operation_facts(is_contribution, contributor_person_key, updated_at DESC);
CREATE TABLE IF NOT EXISTS newsflash_daily_modes (
    duty_date text PRIMARY KEY,
    mode text NOT NULL CHECK (mode IN ('two','three')),
    updated_by text,
    created_at text NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at text NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS newsflash_shift_assignments (
    duty_date text NOT NULL,
    shift_key text NOT NULL CHECK (shift_key IN ('morning','middle','late')),
    person_key text REFERENCES newsflash_roster(person_key) ON DELETE SET NULL,
    updated_by text,
    created_at text NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at text NOT NULL DEFAULT CURRENT_TIMESTAMP,
    PRIMARY KEY (duty_date, shift_key)
);
CREATE INDEX IF NOT EXISTS idx_newsflash_shift_person_date
ON newsflash_shift_assignments(person_key, duty_date);
CREATE TABLE IF NOT EXISTS newsflash_reporting_weeks (
    week_start text PRIMARY KEY,
    report_month text NOT NULL,
    updated_by text,
    created_at text NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at text NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS newsflash_operation_audit (
    id integer PRIMARY KEY AUTOINCREMENT,
    source_item_id text,
    entity_type text NOT NULL,
    entity_key text NOT NULL,
    action text NOT NULL,
    before_json text NOT NULL DEFAULT '{}',
    after_json text NOT NULL DEFAULT '{}',
    actor_email text NOT NULL,
    created_at text NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_newsflash_operation_audit_entity
ON newsflash_operation_audit(entity_type, entity_key, created_at DESC);
CREATE TABLE IF NOT EXISTS newsflash_quality_week_rules (
    week_start text PRIMARY KEY,
    rules_json text NOT NULL,
    created_at text NOT NULL DEFAULT CURRENT_TIMESTAMP
);
"""


DEFAULT_PEOPLE = (
    ("zoey", "Zoey", 1, 1, ("Z", "Zoey")),
    ("harbour", "Harbour", 1, 1, ("蔡", "蔡聪", "Harbour")),
    ("shark", "Shark", 1, 1, ("shark", "Shark")),
    ("leo", "Leo", 1, 1, ("LEO", "CryptoLeo", "Leo")),
    ("malvo", "Malvo", 1, 1, ("南枳", "Malvo")),
    ("golem", "Golem", 0, 1, ("Golem",)),
    ("wenser", "wenser", 0, 1, ("wenser",)),
    ("asher", "asher", 0, 1, ("Asher", "asher")),
    ("azuma", "azuma", 0, 1, ("azuma",)),
    ("dc", "Dc.", 0, 1, ("jk", "Dc", "Dc.")),
)


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)


def _decode(value: Any, default: Any = None) -> Any:
    if value in (None, ""):
        return default
    try:
        return json.loads(str(value))
    except (TypeError, json.JSONDecodeError):
        return default


def _normalize_alias(value: str | None) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def _parse_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value).strip().replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(text)
        except ValueError:
            return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=SHANGHAI_TZ)
    return parsed.astimezone(SHANGHAI_TZ)


def _iso(value: datetime | None = None) -> str:
    return (value or datetime.now(UTC)).isoformat()


def _as_source_item_id(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value or "").strip()
    return text[:-2] if text.endswith(".0") and text[:-2].isdigit() else text


def _week_start(value: date) -> date:
    return value - timedelta(days=value.weekday())


def _month_dates(month_key: str) -> list[date]:
    year, month = (int(part) for part in month_key.split("-", 1))
    return [date(year, month, day) for day in range(1, calendar.monthrange(year, month)[1] + 1)]


def _source_labels(raw_sources: Iterable[str]) -> list[str]:
    order = {key: index for index, key in enumerate(("odaily", "blockbeats", "panews", "jinse"))}
    sources = sorted(set(raw_sources), key=lambda item: (order.get(item, 99), item))
    return [SOURCE_LABELS.get(source, source) for source in sources]


@dataclass(frozen=True, slots=True)
class ShiftWindow:
    duty_date: date
    shift_key: str
    shift_label: str
    person_key: str
    person_name: str
    core_start: datetime
    core_end: datetime

    @property
    def expanded_start(self) -> datetime:
        return self.core_start - timedelta(minutes=30)

    @property
    def expanded_end(self) -> datetime:
        return self.core_end + timedelta(minutes=30)


class NewsflashOperationsRepository:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.init_schema()

    def init_schema(self) -> None:
        with connect_sqlite(self.path) as conn:
            conn.executescript(SCHEMA_SQL)
            if self._table_exists(conn, "newsflash_events"):
                event_columns = {row["name"] for row in conn.execute("PRAGMA table_info(newsflash_events)").fetchall()}
                if "first_sources" not in event_columns:
                    conn.execute("ALTER TABLE newsflash_events ADD COLUMN first_sources text NOT NULL DEFAULT '[]'")
            fact_columns = {row["name"] for row in conn.execute("PRAGMA table_info(newsflash_operation_facts)").fetchall()}
            if "quality_override" not in fact_columns:
                conn.execute("ALTER TABLE newsflash_operation_facts ADD COLUMN quality_override text NOT NULL DEFAULT 'none'")
            conn.execute(
                "CREATE TABLE IF NOT EXISTS newsflash_event_exclusions("
                "source text NOT NULL,source_item_id text NOT NULL,title text,matched_at text NOT NULL DEFAULT CURRENT_TIMESTAMP,"
                "PRIMARY KEY(source,source_item_id))"
            )
            for person_key, display_name, duty_enabled, contributor_enabled, aliases in DEFAULT_PEOPLE:
                conn.execute(
                    "INSERT INTO newsflash_roster(person_key,display_name,duty_enabled,contributor_enabled,active) "
                    "VALUES (?,?,?,?,1) ON CONFLICT(person_key) DO NOTHING",
                    (person_key, display_name, duty_enabled, contributor_enabled),
                )
                for alias in aliases:
                    conn.execute(
                        "INSERT INTO newsflash_roster_aliases(alias_normalized,alias_display,person_key) VALUES (?,?,?) "
                        "ON CONFLICT(alias_normalized) DO UPDATE SET person_key=excluded.person_key,alias_display=excluded.alias_display",
                        (_normalize_alias(alias), alias, person_key),
                    )
            conn.commit()

    def execute(self, action: str, payload: dict[str, Any], *, actor_email: str) -> Any:
        handlers = {
            "list": self.list_newsflashes,
            "update": lambda data: self.update_newsflash(data, actor_email=actor_email),
            "roster": lambda data: self.list_roster(),
            "save_person": lambda data: self.save_person(data, actor_email=actor_email),
            "schedule": self.get_schedule,
            "save_day_mode": lambda data: self.save_day_mode(data, actor_email=actor_email),
            "save_assignment": lambda data: self.save_assignment(data, actor_email=actor_email),
            "save_week_month": lambda data: self.save_week_month(data, actor_email=actor_email),
            "summary": self.get_summary,
            "quality": self.get_quality,
            "contributions": self.list_contributions,
            "events": self.list_events,
        }
        handler = handlers.get(action)
        if handler is None:
            raise ValueError("unsupported newsflash operations action")
        return handler(payload)

    def _audit(
        self,
        conn,
        *,
        source_item_id: str | None,
        entity_type: str,
        entity_key: str,
        action: str,
        before: Any,
        after: Any,
        actor_email: str,
    ) -> None:
        conn.execute(
            "INSERT INTO newsflash_operation_audit(source_item_id,entity_type,entity_key,action,before_json,after_json,actor_email) "
            "VALUES (?,?,?,?,?,?,?)",
            (source_item_id, entity_type, entity_key, action, _json(before or {}), _json(after or {}), actor_email),
        )

    def _person_for_alias(self, conn, operator_raw: str | None) -> str | None:
        normalized = _normalize_alias(operator_raw)
        if not normalized:
            return None
        row = conn.execute(
            "SELECT person_key FROM newsflash_roster_aliases WHERE alias_normalized=?",
            (normalized,),
        ).fetchone()
        return str(row["person_key"]) if row else None

    def _classify_operator(self, conn, operator_raw: str | None) -> tuple[str, str | None]:
        if str(operator_raw or "").strip():
            person_key = self._person_for_alias(conn, operator_raw)
            return ("human" if person_key else "human_unmapped", person_key)
        return ("pending_ai", None)

    def upsert_source_facts(
        self,
        facts: list[dict[str, Any]],
        *,
        snapshot_at: str | None = None,
    ) -> dict[str, int]:
        matched = 0
        skipped = 0
        with connect_sqlite(self.path) as conn:
            for fact in facts:
                source_item_id = _as_source_item_id(fact.get("source_item_id"))
                if not source_item_id:
                    skipped += 1
                    continue
                reference = conn.execute(
                    "SELECT source_item_id FROM odaily_reference_items WHERE source_item_id=?",
                    (source_item_id,),
                ).fetchone()
                if reference is None:
                    skipped += 1
                    continue
                existing = conn.execute(
                    "SELECT publisher_locked FROM newsflash_operation_facts WHERE source_item_id=?",
                    (source_item_id,),
                ).fetchone()
                publisher_kind, publisher_person_key = self._classify_operator(conn, fact.get("operator_raw"))
                if existing and existing["publisher_locked"]:
                    publisher_kind = None
                    publisher_person_key = None
                conn.execute(
                    """
                    INSERT INTO newsflash_operation_facts(
                        source_item_id,operator_raw,publisher_kind,publisher_person_key,view_count,is_pushed,
                        pushed_at,source_snapshot_at
                    ) VALUES (?,?,?,?,?,?,?,?)
                    ON CONFLICT(source_item_id) DO UPDATE SET
                        operator_raw=excluded.operator_raw,
                        publisher_kind=CASE WHEN newsflash_operation_facts.publisher_locked=1 THEN newsflash_operation_facts.publisher_kind ELSE excluded.publisher_kind END,
                        publisher_person_key=CASE WHEN newsflash_operation_facts.publisher_locked=1 THEN newsflash_operation_facts.publisher_person_key ELSE excluded.publisher_person_key END,
                        view_count=excluded.view_count,is_pushed=excluded.is_pushed,pushed_at=excluded.pushed_at,
                        source_snapshot_at=excluded.source_snapshot_at,updated_at=CURRENT_TIMESTAMP
                    """,
                    (
                        source_item_id,
                        str(fact.get("operator_raw") or "").strip() or None,
                        publisher_kind,
                        publisher_person_key,
                        fact.get("view_count"),
                        fact.get("is_pushed"),
                        fact.get("pushed_at"),
                        snapshot_at or _iso(),
                    ),
                )
                conn.execute(
                    "UPDATE odaily_reference_items SET source_url=COALESCE(?,source_url),title=COALESCE(?,title),"
                    "published_at=COALESCE(?,published_at),updated_at=CURRENT_TIMESTAMP WHERE source_item_id=?",
                    (fact.get("source_url"), fact.get("title"), fact.get("published_at"), source_item_id),
                )
                matched += 1
            conn.commit()
        reconciled = self.reconcile_ai_publishers()
        return {"read": len(facts), "matched": matched, "skipped": skipped, "reconciled_odaily": reconciled}

    def reconcile_ai_publishers(self) -> int:
        now = datetime.now(SHANGHAI_TZ)
        updated = 0
        with connect_sqlite(self.path) as conn:
            rows = conn.execute(
                """
                SELECT f.source_item_id,r.title,r.published_at
                FROM newsflash_operation_facts f
                JOIN odaily_reference_items r ON r.source_item_id=f.source_item_id
                WHERE f.publisher_locked=0 AND (f.operator_raw IS NULL OR trim(f.operator_raw)='')
                  AND f.publisher_kind IN ('pending_ai','other_ai')
                """
            ).fetchall()
            for row in rows:
                published = _parse_datetime(row["published_at"])
                if published is None:
                    continue
                match = self._match_odaily_task(conn, row["source_item_id"], row["title"], published)
                if match:
                    kind = "odaily_ai"
                elif now - published >= timedelta(minutes=10):
                    kind = "other_ai"
                else:
                    kind = "pending_ai"
                conn.execute(
                    "UPDATE newsflash_operation_facts SET publisher_kind=?,attribution_checked_at=?,updated_at=CURRENT_TIMESTAMP "
                    "WHERE source_item_id=? AND publisher_locked=0",
                    (kind, _iso(), row["source_item_id"]),
                )
                if kind == "odaily_ai":
                    updated += 1
            conn.commit()
        return updated

    def _match_odaily_task(self, conn, source_item_id: str, title: str | None, published: datetime) -> bool:
        if not self._table_exists(conn, "x_task_pipeline"):
            return False
        candidates = conn.execute(
            """
            SELECT p.final_title,p.publish_completed_at,p.push_result
            FROM x_task_pipeline p
            WHERE p.publish_completed_at IS NOT NULL
              AND p.final_title=?
            """,
            (title,),
        ).fetchall()
        for candidate in candidates:
            if source_item_id in self._remote_ids_from_push_result(candidate["push_result"]):
                return True
            completed = _parse_datetime(candidate["publish_completed_at"])
            if completed and abs((completed - published).total_seconds()) <= 300:
                return True
        return False

    def _remote_ids_from_push_result(self, raw: Any) -> set[str]:
        payload = _decode(raw, {})
        found: set[str] = set()

        def walk(value: Any, key: str = "") -> None:
            if isinstance(value, dict):
                for child_key, child in value.items():
                    walk(child, str(child_key))
                return
            if isinstance(value, list):
                for child in value:
                    walk(child, key)
                return
            if key.casefold() in {"id", "newsflashid", "newsflash_id", "source_item_id"} and value is not None:
                found.add(_as_source_item_id(value))
            if key == "response_text" and isinstance(value, str):
                nested = _decode(value)
                if nested is not None:
                    walk(nested)

        walk(payload)
        return found

    def list_roster(self) -> dict[str, Any]:
        with connect_sqlite(self.path) as conn:
            people = [dict(row) for row in conn.execute("SELECT * FROM newsflash_roster ORDER BY display_name COLLATE NOCASE").fetchall()]
            aliases = conn.execute(
                "SELECT alias_display,person_key FROM newsflash_roster_aliases ORDER BY alias_display COLLATE NOCASE"
            ).fetchall()
        by_person: dict[str, list[str]] = {}
        for alias in aliases:
            by_person.setdefault(str(alias["person_key"]), []).append(str(alias["alias_display"]))
        for person in people:
            person["aliases"] = by_person.get(str(person["person_key"]), [])
            person["duty_enabled"] = bool(person["duty_enabled"])
            person["contributor_enabled"] = bool(person["contributor_enabled"])
            person["active"] = bool(person["active"])
        return {"people": people}

    def save_person(self, payload: dict[str, Any], *, actor_email: str) -> dict[str, Any]:
        person_key = str(payload.get("person_key") or "").strip().casefold()
        display_name = str(payload.get("display_name") or "").strip()
        if not person_key or not display_name:
            raise ValueError("person_key and display_name are required")
        aliases = [str(value).strip() for value in payload.get("aliases") or [] if str(value).strip()]
        with connect_sqlite(self.path) as conn:
            before_row = conn.execute("SELECT * FROM newsflash_roster WHERE person_key=?", (person_key,)).fetchone()
            before = dict(before_row) if before_row else None
            conn.execute(
                """
                INSERT INTO newsflash_roster(person_key,display_name,duty_enabled,contributor_enabled,active)
                VALUES (?,?,?,?,?) ON CONFLICT(person_key) DO UPDATE SET display_name=excluded.display_name,
                duty_enabled=excluded.duty_enabled,contributor_enabled=excluded.contributor_enabled,
                active=excluded.active,updated_at=CURRENT_TIMESTAMP
                """,
                (
                    person_key,
                    display_name,
                    int(bool(payload.get("duty_enabled"))),
                    int(bool(payload.get("contributor_enabled"))),
                    int(payload.get("active", True) is not False),
                ),
            )
            conn.execute("DELETE FROM newsflash_roster_aliases WHERE person_key=?", (person_key,))
            for alias in {display_name, *aliases}:
                conn.execute(
                    "INSERT INTO newsflash_roster_aliases(alias_normalized,alias_display,person_key) VALUES (?,?,?) "
                    "ON CONFLICT(alias_normalized) DO UPDATE SET alias_display=excluded.alias_display,person_key=excluded.person_key",
                    (_normalize_alias(alias), alias, person_key),
                )
            after = dict(conn.execute("SELECT * FROM newsflash_roster WHERE person_key=?", (person_key,)).fetchone())
            self._audit(
                conn,
                source_item_id=None,
                entity_type="person",
                entity_key=person_key,
                action="save",
                before=before,
                after={**after, "aliases": aliases},
                actor_email=actor_email,
            )
            conn.commit()
        return self.list_roster()

    def _event_info(self, conn, source_item_id: str) -> dict[str, Any]:
        excluded = conn.execute(
            "SELECT 1 FROM newsflash_event_exclusions WHERE source='odaily' AND source_item_id=?",
            (source_item_id,),
        ).fetchone() if self._table_exists(conn, "newsflash_event_exclusions") else None
        if excluded:
            return {"status": "excluded", "sources": [], "label": "已排除"}
        event = conn.execute(
            """
            SELECT e.* FROM newsflash_items i
            JOIN newsflash_event_sources s ON s.item_id=i.id
            JOIN newsflash_events e ON e.event_id=s.event_id
            WHERE i.source='odaily' AND i.source_item_id=? AND e.status='active'
            LIMIT 1
            """,
            (source_item_id,),
        ).fetchone() if self._table_exists(conn, "newsflash_events") else None
        if not event:
            return {"status": "unmatched", "sources": [], "label": "未匹配到事件"}
        first_sources = _decode(event["first_sources"], []) if "first_sources" in event.keys() else []
        if not first_sources and event["first_source"]:
            first_sources = [event["first_source"]]
        labels = _source_labels(first_sources)
        if not event["first_published_at"]:
            return {"status": "insufficient", "sources": [], "label": "数据不足", "event_id": event["event_id"]}
        return {
            "status": "ready",
            "sources": labels,
            "label": "、".join(labels),
            "event_id": event["event_id"],
            "published_at": event["first_published_at"],
        }

    @staticmethod
    def _table_exists(conn, table: str) -> bool:
        return conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone() is not None

    def list_newsflashes(self, payload: dict[str, Any]) -> dict[str, Any]:
        page = max(1, int(payload.get("page") or 1))
        page_size = 50
        clauses = ["1=1"]
        params: list[Any] = []
        search = str(payload.get("search") or "").strip()
        if search:
            if search.isdigit():
                clauses.append("r.source_item_id=?")
                params.append(search)
            else:
                clauses.append("(lower(r.title) LIKE ? OR r.source_item_id=?)")
                params.extend([f"%{search.casefold()}%", search])
        for key, column in (("date_from", "r.published_at"), ("date_to", "r.published_at")):
            value = str(payload.get(key) or "").strip()
            if value:
                clauses.append(f"datetime({column}) {'>=' if key == 'date_from' else '<'} datetime(?)")
                params.append(value if key == "date_from" else (date.fromisoformat(value) + timedelta(days=1)).isoformat())
        filters = {
            "publisher_kind": "f.publisher_kind",
            "publisher_person_key": "f.publisher_person_key",
            "contributor_person_key": "f.contributor_person_key",
            "contribution_type": "f.contribution_type",
        }
        for key, column in filters.items():
            value = str(payload.get(key) or "").strip()
            if value:
                clauses.append(f"{column}=?")
                params.append(value)
        if payload.get("is_pushed") in (True, False, 1, 0, "1", "0"):
            clauses.append("f.is_pushed=?")
            params.append(int(str(payload.get("is_pushed")) in {"True", "true", "1"}))
        if payload.get("is_contribution") in (True, False, 1, 0, "1", "0"):
            clauses.append("f.is_contribution=?")
            params.append(int(str(payload.get("is_contribution")) in {"True", "true", "1"}))
        with connect_sqlite(self.path) as conn:
            local_clauses = list(clauses)
            local_params = list(params)
            first_status = str(payload.get("first_status") or "").strip()
            has_exclusions = self._table_exists(conn, "newsflash_event_exclusions")
            has_events = self._table_exists(conn, "newsflash_events")
            exclusion_exists = "EXISTS(SELECT 1 FROM newsflash_event_exclusions x WHERE x.source='odaily' AND x.source_item_id=r.source_item_id)"
            event_exists = "EXISTS(SELECT 1 FROM newsflash_items ni JOIN newsflash_event_sources nes ON nes.item_id=ni.id JOIN newsflash_events ne ON ne.event_id=nes.event_id WHERE ni.source='odaily' AND ni.source_item_id=r.source_item_id AND ne.status='active')"
            if first_status == "excluded" and has_exclusions:
                local_clauses.append(exclusion_exists)
            elif first_status == "unmatched" and has_events:
                if has_exclusions:
                    local_clauses.append(f"NOT {exclusion_exists}")
                local_clauses.append(f"NOT {event_exists}")
            elif first_status in {"ready", "insufficient"} and has_events:
                local_clauses.append(
                    "EXISTS(SELECT 1 FROM newsflash_items ni JOIN newsflash_event_sources nes ON nes.item_id=ni.id "
                    "JOIN newsflash_events ne ON ne.event_id=nes.event_id WHERE ni.source='odaily' AND ni.source_item_id=r.source_item_id "
                    f"AND ne.status='active' AND ne.first_published_at IS {'NOT NULL' if first_status == 'ready' else 'NULL'})"
                )
            first_source = str(payload.get("first_source") or "").strip()
            if first_source and has_events:
                local_clauses.append(
                    "EXISTS(SELECT 1 FROM newsflash_items ni JOIN newsflash_event_sources nes ON nes.item_id=ni.id "
                    "JOIN newsflash_events ne ON ne.event_id=nes.event_id WHERE ni.source='odaily' AND ni.source_item_id=r.source_item_id "
                    "AND ne.status='active' AND (ne.first_source=? OR ne.first_sources LIKE ?))"
                )
                local_params.extend([first_source, f'%"{first_source}"%'])
            where = " AND ".join(local_clauses)
            total = int(conn.execute(
                f"SELECT count(*) FROM odaily_reference_items r LEFT JOIN newsflash_operation_facts f ON f.source_item_id=r.source_item_id WHERE {where}",
                local_params,
            ).fetchone()[0])
            rows = conn.execute(
                f"""
                SELECT r.source_item_id,r.source_url,r.title,r.published_at,r.updated_at AS reference_updated_at,
                       f.operator_raw,f.publisher_kind,f.publisher_person_key,f.publisher_locked,f.view_count,
                       f.is_pushed,f.pushed_at,f.is_contribution,f.contributor_person_key,f.contribution_type,
                       f.quality_override,
                       f.source_snapshot_at,f.updated_at AS operation_updated_at,
                       operator.display_name AS publisher_person_name,contributor.display_name AS contributor_name
                FROM odaily_reference_items r
                LEFT JOIN newsflash_operation_facts f ON f.source_item_id=r.source_item_id
                LEFT JOIN newsflash_roster operator ON operator.person_key=f.publisher_person_key
                LEFT JOIN newsflash_roster contributor ON contributor.person_key=f.contributor_person_key
                WHERE {where}
                ORDER BY datetime(r.published_at) DESC,r.source_item_id DESC
                LIMIT ? OFFSET ?
                """,
                [*local_params, page_size, (page - 1) * page_size],
            ).fetchall()
            data = []
            for row in rows:
                item = dict(row)
                item["publisher_locked"] = bool(item.get("publisher_locked"))
                item["is_contribution"] = bool(item.get("is_contribution"))
                item["is_pushed"] = None if item.get("is_pushed") is None else bool(item["is_pushed"])
                item["quality_override"] = str(item.get("quality_override") or "none")
                item["first_publication"] = self._event_info(conn, str(item["source_item_id"]))
                data.append(item)
        return {"items": data, "page": page, "page_size": page_size, "total": total, "pages": max(1, (total + page_size - 1) // page_size)}

    def update_newsflash(self, payload: dict[str, Any], *, actor_email: str) -> dict[str, Any]:
        source_item_id = _as_source_item_id(payload.get("source_item_id"))
        patch = payload.get("patch") or {}
        if not source_item_id or not isinstance(patch, dict):
            raise ValueError("source_item_id and patch are required")
        allowed = {
            "is_contribution",
            "contributor_person_key",
            "contribution_type",
            "publisher_kind",
            "publisher_person_key",
            "operator_raw",
            "quality_override",
        }
        if set(patch) - allowed:
            raise ValueError("unsupported newsflash field")
        with connect_sqlite(self.path) as conn:
            if not conn.execute("SELECT 1 FROM odaily_reference_items WHERE source_item_id=?", (source_item_id,)).fetchone():
                raise ValueError("newsflash not found")
            conn.execute("INSERT INTO newsflash_operation_facts(source_item_id) VALUES (?) ON CONFLICT(source_item_id) DO NOTHING", (source_item_id,))
            before = dict(conn.execute("SELECT * FROM newsflash_operation_facts WHERE source_item_id=?", (source_item_id,)).fetchone())
            data = dict(patch)
            if "contribution_type" in data and data["contribution_type"] not in CONTRIBUTION_TYPES:
                raise ValueError("invalid contribution type")
            if "quality_override" in data:
                data["quality_override"] = str(data["quality_override"] or "none").strip().casefold()
                if data["quality_override"] not in QUALITY_OVERRIDE_VALUES:
                    raise ValueError("invalid quality override")
            if "is_contribution" in data:
                enabled = bool(data["is_contribution"])
                data["is_contribution"] = int(enabled)
                if not enabled:
                    data["contributor_person_key"] = None
                    data["contribution_type"] = "regular"
                elif not (data.get("contributor_person_key") or before.get("contributor_person_key")):
                    raise ValueError("contributor is required")
            resulting_contribution = bool(data.get("is_contribution", before.get("is_contribution")))
            resulting_kind = str(data.get("publisher_kind", before.get("publisher_kind")) or "")
            if resulting_contribution and resulting_kind in {"odaily_ai", "other_ai", "pending_ai"}:
                raise ValueError("AI newsflash cannot be marked as contribution")
            if "publisher_kind" in data:
                if data["publisher_kind"] not in PUBLISHER_KINDS - {"pending_ai"}:
                    raise ValueError("invalid publisher kind")
                data["publisher_locked"] = 1
                if data["publisher_kind"] != "human":
                    data["publisher_person_key"] = None
                elif not data.get("publisher_person_key"):
                    raise ValueError("human publisher requires a person")
            columns = list(data)
            if columns:
                conn.execute(
                    f"UPDATE newsflash_operation_facts SET {','.join(f'{column}=?' for column in columns)},updated_at=CURRENT_TIMESTAMP WHERE source_item_id=?",
                    [int(value) if isinstance(value, bool) else value for value in data.values()] + [source_item_id],
                )
            after = dict(conn.execute("SELECT * FROM newsflash_operation_facts WHERE source_item_id=?", (source_item_id,)).fetchone())
            self._audit(
                conn,
                source_item_id=source_item_id,
                entity_type="newsflash",
                entity_key=source_item_id,
                action="update",
                before=before,
                after=after,
                actor_email=actor_email,
            )
            conn.commit()
        return self.list_newsflashes({"search": source_item_id})["items"][0]

    def ensure_month(self, month_key: str) -> None:
        dates = _month_dates(month_key)
        with connect_sqlite(self.path) as conn:
            for duty_date in dates:
                mode = "three" if duty_date.weekday() < 5 else "two"
                conn.execute(
                    "INSERT INTO newsflash_daily_modes(duty_date,mode) VALUES (?,?) ON CONFLICT(duty_date) DO NOTHING",
                    (duty_date.isoformat(), mode),
                )
            conn.commit()

    def get_schedule(self, payload: dict[str, Any]) -> dict[str, Any]:
        month_key = str(payload.get("month") or datetime.now(SHANGHAI_TZ).strftime("%Y-%m"))
        self.ensure_month(month_key)
        dates = _month_dates(month_key)
        first_week = _week_start(dates[0])
        last_week = _week_start(dates[-1])
        with connect_sqlite(self.path) as conn:
            modes = {row["duty_date"]: row["mode"] for row in conn.execute(
                "SELECT duty_date,mode FROM newsflash_daily_modes WHERE duty_date>=? AND duty_date<=?",
                (dates[0].isoformat(), dates[-1].isoformat()),
            ).fetchall()}
            assignments = [dict(row) for row in conn.execute(
                "SELECT a.*,p.display_name FROM newsflash_shift_assignments a LEFT JOIN newsflash_roster p ON p.person_key=a.person_key "
                "WHERE a.duty_date>=? AND a.duty_date<=? ORDER BY a.duty_date,a.shift_key",
                (dates[0].isoformat(), dates[-1].isoformat()),
            ).fetchall()]
            reporting = {row["week_start"]: row["report_month"] for row in conn.execute(
                "SELECT week_start,report_month FROM newsflash_reporting_weeks WHERE week_start>=? AND week_start<=?",
                (first_week.isoformat(), last_week.isoformat()),
            ).fetchall()}
        weeks = []
        cursor = first_week
        while cursor <= last_week:
            weeks.append({
                "week_start": cursor.isoformat(),
                "week_end": (cursor + timedelta(days=6)).isoformat(),
                "report_month": reporting.get(cursor.isoformat()) or self._default_report_month(cursor),
                "report_month_manual": cursor.isoformat() in reporting,
            })
            cursor += timedelta(days=7)
        return {
            "month": month_key,
            "days": [{"date": value.isoformat(), "mode": modes[value.isoformat()]} for value in dates],
            "assignments": assignments,
            "weeks": weeks,
            **self.list_roster(),
        }

    def _default_report_month(self, week_start: date) -> str | None:
        week_dates = [week_start + timedelta(days=index) for index in range(7)]
        months = {value.strftime("%Y-%m") for value in week_dates}
        return next(iter(months)) if len(months) == 1 else None

    def save_day_mode(self, payload: dict[str, Any], *, actor_email: str) -> dict[str, Any]:
        duty_date = date.fromisoformat(str(payload.get("date")))
        mode = str(payload.get("mode") or "")
        if mode not in SHIFT_DEFINITIONS:
            raise ValueError("invalid day mode")
        key = duty_date.isoformat()
        with connect_sqlite(self.path) as conn:
            before = conn.execute("SELECT * FROM newsflash_daily_modes WHERE duty_date=?", (key,)).fetchone()
            conn.execute(
                "INSERT INTO newsflash_daily_modes(duty_date,mode,updated_by) VALUES (?,?,?) "
                "ON CONFLICT(duty_date) DO UPDATE SET mode=excluded.mode,updated_by=excluded.updated_by,updated_at=CURRENT_TIMESTAMP",
                (key, mode, actor_email),
            )
            if not before or before["mode"] != mode:
                conn.execute("DELETE FROM newsflash_shift_assignments WHERE duty_date=?", (key,))
            after = dict(conn.execute("SELECT * FROM newsflash_daily_modes WHERE duty_date=?", (key,)).fetchone())
            self._audit(conn, source_item_id=None, entity_type="day_mode", entity_key=key, action="save", before=dict(before) if before else None, after=after, actor_email=actor_email)
            conn.commit()
        return self.get_schedule({"month": duty_date.strftime("%Y-%m")})

    def save_assignment(self, payload: dict[str, Any], *, actor_email: str) -> dict[str, Any]:
        duty_date = date.fromisoformat(str(payload.get("date")))
        shift_key = str(payload.get("shift_key") or "")
        person_key = str(payload.get("person_key") or "").strip() or None
        key = duty_date.isoformat()
        with connect_sqlite(self.path) as conn:
            mode_row = conn.execute("SELECT mode FROM newsflash_daily_modes WHERE duty_date=?", (key,)).fetchone()
            mode = str(mode_row["mode"]) if mode_row else ("three" if duty_date.weekday() < 5 else "two")
            valid_shifts = {definition[0] for definition in SHIFT_DEFINITIONS[mode]}
            if shift_key not in valid_shifts:
                raise ValueError("shift is not available for this day mode")
            if person_key and not conn.execute("SELECT 1 FROM newsflash_roster WHERE person_key=? AND active=1 AND duty_enabled=1", (person_key,)).fetchone():
                raise ValueError("duty person is not active")
            before = conn.execute("SELECT * FROM newsflash_shift_assignments WHERE duty_date=? AND shift_key=?", (key, shift_key)).fetchone()
            conn.execute(
                "INSERT INTO newsflash_shift_assignments(duty_date,shift_key,person_key,updated_by) VALUES (?,?,?,?) "
                "ON CONFLICT(duty_date,shift_key) DO UPDATE SET person_key=excluded.person_key,updated_by=excluded.updated_by,updated_at=CURRENT_TIMESTAMP",
                (key, shift_key, person_key, actor_email),
            )
            after = dict(conn.execute("SELECT * FROM newsflash_shift_assignments WHERE duty_date=? AND shift_key=?", (key, shift_key)).fetchone())
            self._audit(conn, source_item_id=None, entity_type="assignment", entity_key=f"{key}:{shift_key}", action="save", before=dict(before) if before else None, after=after, actor_email=actor_email)
            conn.commit()
        return self.get_schedule({"month": duty_date.strftime("%Y-%m")})

    def save_week_month(self, payload: dict[str, Any], *, actor_email: str) -> dict[str, Any]:
        week_start = _week_start(date.fromisoformat(str(payload.get("week_start"))))
        report_month = str(payload.get("report_month") or "")
        covered = {(week_start + timedelta(days=index)).strftime("%Y-%m") for index in range(7)}
        if report_month not in covered:
            raise ValueError("report month must be covered by the natural week")
        key = week_start.isoformat()
        with connect_sqlite(self.path) as conn:
            before = conn.execute("SELECT * FROM newsflash_reporting_weeks WHERE week_start=?", (key,)).fetchone()
            conn.execute(
                "INSERT INTO newsflash_reporting_weeks(week_start,report_month,updated_by) VALUES (?,?,?) "
                "ON CONFLICT(week_start) DO UPDATE SET report_month=excluded.report_month,updated_by=excluded.updated_by,updated_at=CURRENT_TIMESTAMP",
                (key, report_month, actor_email),
            )
            after = dict(conn.execute("SELECT * FROM newsflash_reporting_weeks WHERE week_start=?", (key,)).fetchone())
            self._audit(conn, source_item_id=None, entity_type="reporting_week", entity_key=key, action="save", before=dict(before) if before else None, after=after, actor_email=actor_email)
            conn.commit()
        return {"week_start": key, "report_month": report_month}

    def _period_dates(self, conn, payload: dict[str, Any]) -> tuple[list[date], str, list[str]]:
        report_month = str(payload.get("report_month") or "").strip()
        if report_month:
            weeks = {str(row["week_start"]) for row in conn.execute(
                "SELECT week_start FROM newsflash_reporting_weeks WHERE report_month=? ORDER BY week_start",
                (report_month,),
            ).fetchall()}
            month_dates = _month_dates(report_month)
            cursor = _week_start(month_dates[0])
            last = _week_start(month_dates[-1])
            while cursor <= last:
                covered = {cursor + timedelta(days=index) for index in range(7)}
                if all(value.strftime("%Y-%m") == report_month for value in covered):
                    weeks.add(cursor.isoformat())
                cursor += timedelta(days=7)
            ordered_weeks = sorted(weeks)
            dates = [date.fromisoformat(week) + timedelta(days=index) for week in ordered_weeks for index in range(7)]
            return dates, report_month, ordered_weeks
        start = _week_start(date.fromisoformat(str(payload.get("week_start") or _week_start(datetime.now(SHANGHAI_TZ).date()))))
        return [start + timedelta(days=index) for index in range(7)], start.isoformat(), [start.isoformat()]

    def _shift_windows(self, conn, dates: list[date]) -> list[ShiftWindow]:
        if not dates:
            return []
        start, end = min(dates).isoformat(), max(dates).isoformat()
        modes = {row["duty_date"]: row["mode"] for row in conn.execute(
            "SELECT duty_date,mode FROM newsflash_daily_modes WHERE duty_date>=? AND duty_date<=?",
            (start, end),
        ).fetchall()}
        assignments = conn.execute(
            """
            SELECT a.duty_date,a.shift_key,a.person_key,p.display_name
            FROM newsflash_shift_assignments a JOIN newsflash_roster p ON p.person_key=a.person_key
            WHERE a.duty_date>=? AND a.duty_date<=? AND a.person_key IS NOT NULL
            """,
            (start, end),
        ).fetchall()
        windows = []
        for assignment in assignments:
            duty_date = date.fromisoformat(assignment["duty_date"])
            mode = modes.get(assignment["duty_date"], "three" if duty_date.weekday() < 5 else "two")
            definition = next((item for item in SHIFT_DEFINITIONS[mode] if item[0] == assignment["shift_key"]), None)
            if definition is None:
                continue
            shift_key, label, start_time, end_time, end_offset = definition
            windows.append(ShiftWindow(
                duty_date=duty_date,
                shift_key=shift_key,
                shift_label=label,
                person_key=assignment["person_key"],
                person_name=assignment["display_name"],
                core_start=datetime.combine(duty_date, start_time, SHANGHAI_TZ),
                core_end=datetime.combine(duty_date + timedelta(days=end_offset), end_time, SHANGHAI_TZ),
            ))
        return windows

    @staticmethod
    def _window_distance(moment: datetime, window: ShiftWindow) -> float:
        if window.core_start <= moment < window.core_end:
            return 0.0
        if moment < window.core_start:
            return (window.core_start - moment).total_seconds()
        return (moment - window.core_end).total_seconds()

    def _assign_window(self, moment: datetime, person_key: str, windows: list[ShiftWindow]) -> ShiftWindow | None:
        candidates = [window for window in windows if window.person_key == person_key and window.expanded_start <= moment < window.expanded_end]
        if not candidates:
            return None
        core = [window for window in candidates if window.core_start <= moment < window.core_end]
        if core:
            return sorted(core, key=lambda window: window.core_start)[0]
        return min(candidates, key=lambda window: (self._window_distance(moment, window), window.core_start))

    def _metric(self, items: list[dict[str, Any]]) -> dict[str, Any]:
        views = [int(item["view_count"]) for item in items if item.get("view_count") is not None]
        pushed = [item for item in items if item.get("is_pushed") is True]
        push_views = [int(item["view_count"]) for item in pushed if item.get("view_count") is not None]
        push_known = [item for item in items if item.get("is_pushed") is not None]
        return {
            "published_count": len(items),
            "pushed_count": len(pushed),
            "average_views": round(sum(views) / len(views), 1) if views else None,
            "pushed_views": sum(push_views) if push_views else (0 if pushed else None),
            "view_coverage": {"known": len(views), "total": len(items)},
            "push_coverage": {"known": len(push_known), "total": len(items)},
            "push_view_coverage": {"known": len(push_views), "total": len(pushed)},
        }

    def get_summary(self, payload: dict[str, Any]) -> dict[str, Any]:
        with connect_sqlite(self.path) as conn:
            dates, period_label, weeks = self._period_dates(conn, payload)
            if not dates:
                return {"period": period_label, "weeks": weeks, "rows": [], "people": [], "unassigned_count": 0}
            windows = self._shift_windows(conn, dates)
            period_start = datetime.combine(min(dates), time(0, 0), SHANGHAI_TZ) - timedelta(minutes=30)
            period_end = datetime.combine(max(dates) + timedelta(days=2), time(0, 0), SHANGHAI_TZ)
            source_rows = conn.execute(
                """
                SELECT r.source_item_id,r.published_at,f.publisher_kind,f.publisher_person_key,f.view_count,
                       f.is_pushed,f.is_contribution
                FROM odaily_reference_items r JOIN newsflash_operation_facts f ON f.source_item_id=r.source_item_id
                WHERE datetime(r.published_at)>=datetime(?) AND datetime(r.published_at)<datetime(?)
                """,
                (period_start.isoformat(), period_end.isoformat()),
            ).fetchall()
            roster = {row["person_key"]: row["display_name"] for row in conn.execute("SELECT person_key,display_name FROM newsflash_roster").fetchall()}
        assigned: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
        unassigned = 0
        ai_by_day: dict[str, list[dict[str, Any]]] = {}
        date_set = set(dates)
        for row in source_rows:
            item = dict(row)
            item["is_pushed"] = None if item["is_pushed"] is None else bool(item["is_pushed"])
            moment = _parse_datetime(item["published_at"])
            if moment is None:
                continue
            if item["publisher_kind"] == "odaily_ai" and moment.date() in date_set:
                ai_by_day.setdefault(moment.date().isoformat(), []).append(item)
                continue
            if item["publisher_kind"] == "human_unmapped":
                if moment.date() in date_set:
                    unassigned += 1
                continue
            if item["publisher_kind"] != "human" or item["is_contribution"]:
                continue
            window = self._assign_window(moment, str(item["publisher_person_key"] or ""), windows)
            if window is None:
                if moment.date() in date_set:
                    unassigned += 1
                continue
            key = (window.duty_date.isoformat(), window.shift_key, window.person_key)
            assigned.setdefault(key, []).append(item)
        rows = []
        for window in sorted(windows, key=lambda item: (item.duty_date, item.core_start, item.person_name)):
            key = (window.duty_date.isoformat(), window.shift_key, window.person_key)
            rows.append({
                "date": key[0],
                "shift_key": window.shift_key,
                "shift_label": window.shift_label,
                "person_key": window.person_key,
                "person_name": window.person_name,
                "is_ai": False,
                **self._metric(assigned.get(key, [])),
            })
        for duty_date in sorted(date_set):
            rows.append({
                "date": duty_date.isoformat(),
                "shift_key": "all_day",
                "shift_label": "全天",
                "person_key": "odaily_ai",
                "person_name": "OdAIly",
                "is_ai": True,
                **self._metric(ai_by_day.get(duty_date.isoformat(), [])),
            })
        grouped: dict[str, list[dict[str, Any]]] = {}
        for row in rows:
            grouped.setdefault(row["person_key"], []).append(row)
        people = []
        for person_key, person_rows in grouped.items():
            aggregate_items = []
            # Recreate aggregate metrics from row totals while preserving weighted average coverage.
            published = sum(int(row["published_count"]) for row in person_rows)
            pushed = sum(int(row["pushed_count"]) for row in person_rows)
            view_known = sum(int(row["view_coverage"]["known"]) for row in person_rows)
            view_total = sum(int(row["view_coverage"]["total"]) for row in person_rows)
            weighted_views = sum((row["average_views"] or 0) * row["view_coverage"]["known"] for row in person_rows)
            pushed_views = sum(int(row["pushed_views"] or 0) for row in person_rows)
            people.append({
                "person_key": person_key,
                "person_name": "OdAIly" if person_key == "odaily_ai" else roster.get(person_key, person_key),
                "published_count": published,
                "pushed_count": pushed,
                "average_views": round(weighted_views / view_known, 1) if view_known else None,
                "pushed_views": pushed_views if pushed else None,
                "view_coverage": {"known": view_known, "total": view_total},
            })
        return {"period": period_label, "weeks": weeks, "rows": rows, "people": people, "unassigned_count": unassigned}

    @staticmethod
    def _url_host(value: str | None) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        try:
            host = (urlparse(text if "://" in text else f"https://{text}").hostname or "").casefold()
        except ValueError:
            return ""
        return host.removeprefix("www.")

    @classmethod
    def _is_odaily_url(cls, value: str | None) -> bool:
        host = cls._url_host(value)
        return host == "odaily.news" or host.endswith(".odaily.news")

    @classmethod
    def _x_username(cls, value: str | None) -> str | None:
        text = str(value or "").strip()
        if not text:
            return None
        try:
            parsed = urlparse(text if "://" in text else f"https://{text}")
        except ValueError:
            return None
        host = (parsed.hostname or "").casefold().removeprefix("www.").removeprefix("mobile.")
        if host not in {"x.com", "twitter.com"}:
            return None
        segment = next((part for part in parsed.path.split("/") if part), "")
        return segment.removeprefix("@").casefold() or None

    @staticmethod
    def _host_matches(host: str, domains: Iterable[str]) -> bool:
        return bool(host) and any(host == domain or host.endswith(f".{domain}") for domain in domains)

    def _quality_original_url(self, row: Any) -> str | None:
        raw_payload = _decode(row["raw_payload"], {})
        raw_source = raw_payload.get("sourceUrl") if isinstance(raw_payload, dict) else None
        if str(raw_source or "").strip():
            return str(raw_source).strip()
        fallback = str(row["source_url"] or "").strip()
        return fallback if fallback and not self._is_odaily_url(fallback) else None

    def _quality_first_publication(self, conn, source_item_id: str) -> tuple[dict[str, Any], list[str]]:
        info = self._event_info(conn, source_item_id)
        if info.get("status") != "ready" or not self._table_exists(conn, "newsflash_events"):
            return info, []
        event = conn.execute(
            """
            SELECT e.first_source,e.first_sources FROM newsflash_items i
            JOIN newsflash_event_sources s ON s.item_id=i.id
            JOIN newsflash_events e ON e.event_id=s.event_id
            WHERE i.source='odaily' AND i.source_item_id=? AND e.status='active' LIMIT 1
            """,
            (source_item_id,),
        ).fetchone()
        if event is None:
            return info, []
        sources = _decode(event["first_sources"], []) or ([event["first_source"]] if event["first_source"] else [])
        return info, [str(source).casefold() for source in sources if str(source or "").strip()]

    def _quality_rule_seed(self, conn) -> dict[str, Any]:
        x_accounts: list[str] = []
        if self._table_exists(conn, "x_capture_accounts"):
            x_accounts = [str(row["username"]) for row in conn.execute(
                "SELECT username FROM x_capture_accounts WHERE enabled=1 ORDER BY username_lower,username"
            ).fetchall()]
        media_urls = list(QUALITY_EXTERNAL_MEDIA_URLS)
        if self._table_exists(conn, "non_mainstream_media_sources"):
            media_urls.extend(str(row["homepage_url"]) for row in conn.execute(
                "SELECT homepage_url FROM non_mainstream_media_sources WHERE enabled=1 ORDER BY site_key"
            ).fetchall() if str(row["homepage_url"] or "").strip())
        media_domains = sorted({self._url_host(value) for value in media_urls if self._url_host(value)})
        return {
            "regular_source_accounts": list(QUALITY_REGULAR_SOURCE_ACCOUNTS),
            "automated_x_accounts": x_accounts,
            "automated_media_domains": media_domains,
            "keyword_groups": [
                {
                    "key": group["key"],
                    "terms": list(group["terms"]),
                    "label": " + ".join(group["terms"]),
                }
                for group in QUALITY_EXCLUSION_GROUPS
            ],
            "threshold_multiplier": QUALITY_THRESHOLD_MULTIPLIER,
            "kpi_per_item": QUALITY_KPI_PER_ITEM,
        }

    def _quality_rules(self, conn, week_start: date) -> dict[str, Any]:
        key = week_start.isoformat()
        seed = self._quality_rule_seed(conn)
        conn.execute(
            "INSERT OR IGNORE INTO newsflash_quality_week_rules(week_start,rules_json,created_at) VALUES (?,?,?)",
            (key, _json(seed), _iso()),
        )
        row = conn.execute(
            "SELECT rules_json,created_at FROM newsflash_quality_week_rules WHERE week_start=?",
            (key,),
        ).fetchone()
        rules = _decode(row["rules_json"], {})
        if "keyword_groups" not in rules:
            # Existing week snapshots predate keyword exclusions. Add the new
            # fixed rule set once so subsequent reads remain snapshot-based.
            rules["keyword_groups"] = seed["keyword_groups"]
            conn.execute(
                "UPDATE newsflash_quality_week_rules SET rules_json=? WHERE week_start=?",
                (_json(rules), key),
            )
        conn.commit()
        rules["snapshot_at"] = row["created_at"]
        return rules

    def _latest_quality_week(self, conn) -> date:
        row = conn.execute(
            """
            SELECT max(r.published_at) AS published_at
            FROM newsflash_operation_facts f
            JOIN odaily_reference_items r ON r.source_item_id=f.source_item_id
            WHERE f.is_pushed=1 AND f.view_count IS NOT NULL AND datetime(r.published_at)>=datetime(?)
            """,
            (datetime.combine(QUALITY_FIRST_WEEK, time(0, 0), SHANGHAI_TZ).isoformat(),),
        ).fetchone()
        published = _parse_datetime(row["published_at"] if row else None)
        return _week_start(published.date()) if published else max(QUALITY_FIRST_WEEK, _week_start(datetime.now(SHANGHAI_TZ).date()))

    def get_quality(self, payload: dict[str, Any]) -> dict[str, Any]:
        with connect_sqlite(self.path) as conn:
            raw_week = str(payload.get("week_start") or "").strip()
            start = _week_start(date.fromisoformat(raw_week)) if raw_week else self._latest_quality_week(conn)
            if start < QUALITY_FIRST_WEEK:
                raise ValueError(f"quality week must not be earlier than {QUALITY_FIRST_WEEK.isoformat()}")
            end = start + timedelta(days=7)
            rules = self._quality_rules(conn, start)
            period_start = datetime.combine(start, time(0, 0), SHANGHAI_TZ)
            period_end = datetime.combine(end, time(0, 0), SHANGHAI_TZ)
            rows = conn.execute(
                """
                SELECT r.source_item_id,r.source_url,r.title,r.content,r.raw_payload,r.published_at,
                       f.publisher_kind,f.publisher_person_key,f.view_count,f.is_pushed,f.is_contribution,
                       f.quality_override
                FROM odaily_reference_items r
                JOIN newsflash_operation_facts f ON f.source_item_id=r.source_item_id
                WHERE datetime(r.published_at)>=datetime(?) AND datetime(r.published_at)<datetime(?)
                ORDER BY datetime(r.published_at) DESC,r.source_item_id DESC
                """,
                (period_start.isoformat(), period_end.isoformat()),
            ).fetchall()
            pushed = [row for row in rows if row["is_pushed"] == 1]
            pushed_views = [int(row["view_count"]) for row in pushed if row["view_count"] is not None]
            base = {
                "status": "ready" if pushed_views else "insufficient",
                "week_start": start.isoformat(),
                "week_end": (end - timedelta(days=1)).isoformat(),
                "pushed_count": len(pushed),
                "pushed_view_count": len(pushed_views),
                "rules": rules,
            }
            if not pushed_views:
                return {
                    **base, "average_views": None, "threshold_views": None,
                    "qualified_count": 0, "excluded_count": 0, "total_kpi": 0,
                    "unassigned_count": 0, "groups": [],
                }

            average = sum(pushed_views) / len(pushed_views)
            threshold = average * float(rules["threshold_multiplier"])
            dates = [start + timedelta(days=index) for index in range(7)]
            windows = self._shift_windows(conn, dates)
            people = {
                window.person_key: window.person_name
                for window in sorted(windows, key=lambda value: (value.person_name.casefold(), value.person_key))
            }
            grouped = {key: {"qualified": [], "excluded": []} for key in people}
            unassigned_count = 0
            regular_accounts = {str(value).casefold() for value in rules["regular_source_accounts"]}
            automated_accounts = {str(value).casefold() for value in rules["automated_x_accounts"]}
            automated_domains = {str(value).casefold() for value in rules["automated_media_domains"]}
            keyword_groups = rules.get("keyword_groups", [])

            for row in rows:
                moment = _parse_datetime(row["published_at"])
                if moment is None or row["publisher_kind"] not in {"human", "human_unmapped"}:
                    continue
                window = self._assign_window(moment, str(row["publisher_person_key"] or ""), windows) if row["publisher_kind"] == "human" else None
                if window is None:
                    unassigned_count += 1
                    continue
                quality_override = str(row["quality_override"] or "none").strip().casefold()
                if quality_override != "include" and (row["view_count"] is None or int(row["view_count"]) <= threshold):
                    continue
                reasons: list[str] = []
                reason_labels: list[str] = []
                original_url = self._quality_original_url(row)
                username = self._x_username(original_url)
                host = self._url_host(original_url)
                first_publication, first_sources = self._quality_first_publication(conn, str(row["source_item_id"]))
                if quality_override != "include":
                    if row["is_contribution"]:
                        reasons.append("contribution")
                        reason_labels.append("贡献快讯")
                    if first_publication.get("status") == "ready" and first_sources and "odaily" not in first_sources:
                        reasons.append("competitor_first")
                        reason_labels.append("晚于竞品")
                    if username and username in regular_accounts:
                        reasons.append("regular_source")
                        reason_labels.append("常规信源")
                    if (username and username in automated_accounts) or self._host_matches(host, automated_domains):
                        reasons.append("automated_coverage")
                        reason_labels.append("自动覆盖")
                    if "金十" in str(row["content"] or ""):
                        reasons.append("jin10_content")
                        reason_labels.append("正文含金十")
                    quality_text = f"{row['title'] or ''}\n{row['content'] or ''}".casefold()
                    for group in keyword_groups:
                        terms = [str(term) for term in group.get("terms", []) if str(term).strip()]
                        if terms and all(term.casefold() in quality_text for term in terms):
                            reasons.append(f"keyword_{group.get('key', 'unknown')}")
                            reason_labels.append(f"排除词：{group.get('label') or ' + '.join(terms)}")
                    if quality_override == "exclude":
                        reasons.append("manual_exclude")
                        reason_labels.append("人工排除")
                item = {
                    "source_item_id": str(row["source_item_id"]),
                    "odaily_url": f"https://www.odaily.news/zh-CN/newsflash/{row['source_item_id']}",
                    "original_url": original_url,
                    "has_original_url": bool(original_url),
                    "title": row["title"],
                    "published_at": row["published_at"],
                    "view_count": None if row["view_count"] is None else int(row["view_count"]),
                    "is_pushed": None if row["is_pushed"] is None else bool(row["is_pushed"]),
                    "first_publication": first_publication,
                    "quality_override": quality_override,
                    "exclusion_reasons": reasons,
                    "exclusion_reason_labels": reason_labels,
                }
                grouped[window.person_key]["excluded" if reasons else "qualified"].append(item)

            groups = []
            for person_key, person_name in people.items():
                qualified = grouped[person_key]["qualified"]
                excluded = grouped[person_key]["excluded"]
                groups.append({
                    "person_key": person_key,
                    "person_name": person_name,
                    "qualified_count": len(qualified),
                    "excluded_count": len(excluded),
                    "kpi": round(len(qualified) * float(rules["kpi_per_item"]), 10),
                    "qualified": qualified,
                    "excluded": excluded,
                })
            qualified_count = sum(group["qualified_count"] for group in groups)
            excluded_count = sum(group["excluded_count"] for group in groups)
            return {
                **base,
                "average_views": average,
                "threshold_views": threshold,
                "qualified_count": qualified_count,
                "excluded_count": excluded_count,
                "total_kpi": round(qualified_count * float(rules["kpi_per_item"]), 10),
                "unassigned_count": unassigned_count,
                "groups": groups,
            }

    def list_contributions(self, payload: dict[str, Any]) -> dict[str, Any]:
        start = _week_start(date.fromisoformat(str(payload.get("week_start") or _week_start(datetime.now(SHANGHAI_TZ).date()))))
        end = start + timedelta(days=7)
        with connect_sqlite(self.path) as conn:
            period_start = datetime.combine(start, time(0, 0), SHANGHAI_TZ).isoformat()
            period_end = datetime.combine(end, time(0, 0), SHANGHAI_TZ).isoformat()
            baseline_rows = conn.execute(
                """
                SELECT f.view_count
                FROM newsflash_operation_facts f JOIN odaily_reference_items r ON r.source_item_id=f.source_item_id
                WHERE f.is_pushed=1 AND datetime(r.published_at)>=datetime(?) AND datetime(r.published_at)<datetime(?)
                """,
                (period_start, period_end),
            ).fetchall()
            baseline_views = [int(row["view_count"]) for row in baseline_rows if row["view_count"] is not None]
            average_views = sum(baseline_views) / len(baseline_views) if baseline_views else None
            people = [dict(row) for row in conn.execute(
                "SELECT person_key,display_name FROM newsflash_roster WHERE active=1 AND contributor_enabled=1 ORDER BY display_name COLLATE NOCASE"
            ).fetchall()]
            rows = [dict(row) for row in conn.execute(
                """
                SELECT r.source_item_id,r.source_url,r.title,r.published_at,f.view_count,f.contribution_type,
                       f.contributor_person_key,p.display_name AS contributor_name
                FROM newsflash_operation_facts f JOIN odaily_reference_items r ON r.source_item_id=f.source_item_id
                JOIN newsflash_roster p ON p.person_key=f.contributor_person_key
                WHERE f.is_contribution=1 AND datetime(r.published_at)>=datetime(?) AND datetime(r.published_at)<datetime(?)
                ORDER BY p.display_name COLLATE NOCASE,datetime(r.published_at) DESC
                """,
                (period_start, period_end),
            ).fetchall()]
            for row in rows:
                row["first_publication"] = self._event_info(conn, str(row["source_item_id"]))
                score = self._contribution_score(row["view_count"], row["contribution_type"], average_views)
                row.update(score)
        groups = []
        for person in people:
            items = [row for row in rows if row["contributor_person_key"] == person["person_key"]]
            views = [int(row["view_count"]) for row in items if row["view_count"] is not None]
            base_score_total = round(sum(float(item["base_score"]) for item in items), 10)
            high_view_bonus_count = sum(1 for item in items if float(item["high_view_bonus"]) > 0)
            high_view_bonus = round(sum(float(item["high_view_bonus"]) for item in items), 10)
            base_score_capped = min(base_score_total, CONTRIBUTION_SCORE_CAP)
            groups.append({
                **person,
                "count": len(items),
                "total_views": sum(views),
                "average_views": round(sum(views) / len(views), 1) if views else None,
                "view_coverage": {"known": len(views), "total": len(items)},
                "base_score_total": base_score_total,
                "base_score_capped": base_score_capped,
                "high_view_bonus_count": high_view_bonus_count,
                "high_view_bonus": high_view_bonus,
                "total_score": round(base_score_capped + high_view_bonus, 10),
                "items": items,
            })
        return {
            "week_start": start.isoformat(),
            "week_end": (end - timedelta(days=1)).isoformat(),
            "in_progress": datetime.now(SHANGHAI_TZ).date() < end,
            "status": "ready" if average_views is not None else "insufficient",
            "baseline": {
                "pushed_count": len(baseline_rows),
                "known_view_count": len(baseline_views),
                "average_views": average_views,
            },
            "groups": groups,
        }

    @staticmethod
    def _contribution_score(view_count: Any, contribution_type: str | None, average_views: float | None) -> dict[str, float]:
        if contribution_type in {"night", "ppp"}:
            return {"base_score": 0.5, "high_view_bonus": 0.0, "score": 0.5}
        if view_count is None or average_views is None:
            return {"base_score": 0.0, "high_view_bonus": 0.0, "score": 0.0}
        views = float(view_count)
        if views < average_views * 0.5:
            base_score = 0.0
        elif views < average_views * 0.75:
            base_score = 0.25
        elif views <= average_views * 2:
            base_score = 0.5
        else:
            base_score = 1.0
        high_view_bonus = CONTRIBUTION_HIGH_VIEW_BONUS if views > average_views * 2 else 0.0
        return {
            "base_score": base_score,
            "high_view_bonus": high_view_bonus,
            "score": base_score + high_view_bonus,
        }

    def list_events(self, payload: dict[str, Any]) -> dict[str, Any]:
        page = max(1, int(payload.get("page") or 1))
        page_size = 50
        clauses = ["e.status='active'"]
        params: list[Any] = []
        search = str(payload.get("search") or "").strip()
        if search:
            search_pattern = f"%{search.casefold()}%"
            clauses.append(
                """(
                    lower(COALESCE(e.representative_title, '')) LIKE ?
                    OR EXISTS (
                        SELECT 1
                        FROM newsflash_event_sources search_sources
                        JOIN newsflash_items search_items ON search_items.id=search_sources.item_id
                        WHERE search_sources.event_id=e.event_id
                          AND (
                              lower(COALESCE(search_items.title, '')) LIKE ?
                              OR lower(COALESCE(search_items.source_item_id, '')) LIKE ?
                          )
                    )
                )"""
            )
            params.extend([search_pattern, search_pattern, search_pattern])
        if payload.get("has_odaily") in (True, False, 1, 0, "1", "0"):
            clauses.append("e.has_odaily=?")
            params.append(int(str(payload.get("has_odaily")) in {"True", "true", "1"}))
        first_source = str(payload.get("first_source") or "").strip()
        if first_source:
            clauses.append("(e.first_source=? OR e.first_sources LIKE ?)")
            params.extend([first_source, f'%"{first_source}"%'])
        date_from = str(payload.get("date_from") or "").strip()
        date_to = str(payload.get("date_to") or "").strip()
        if date_from:
            clauses.append("datetime(e.event_time)>=datetime(?)")
            params.append(date_from)
        if date_to:
            clauses.append("datetime(e.event_time)<datetime(?)")
            params.append((date.fromisoformat(date_to) + timedelta(days=1)).isoformat())
        where = " AND ".join(clauses)
        with connect_sqlite(self.path) as conn:
            if not self._table_exists(conn, "newsflash_events"):
                return {"items": [], "page": 1, "pages": 1, "total": 0, "page_size": page_size}
            total = int(conn.execute(f"SELECT count(*) FROM newsflash_events e WHERE {where}", params).fetchone()[0])
            events = conn.execute(
                f"SELECT e.* FROM newsflash_events e WHERE {where} ORDER BY datetime(e.event_time) DESC,e.updated_at DESC LIMIT ? OFFSET ?",
                [*params, page_size, (page - 1) * page_size],
            ).fetchall()
            items = []
            for event in events:
                first_sources = _decode(event["first_sources"], []) if "first_sources" in event.keys() else ([event["first_source"]] if event["first_source"] else [])
                sources = [dict(row) for row in conn.execute(
                    """
                    SELECT i.source,i.source_item_id,i.source_url,i.title,i.published_at
                    FROM newsflash_event_sources s JOIN newsflash_items i ON i.id=s.item_id
                    WHERE s.event_id=? ORDER BY CASE WHEN i.published_at IS NULL THEN 1 ELSE 0 END,datetime(i.published_at),i.id
                    """,
                    (event["event_id"],),
                ).fetchall()]
                odaily = next((source for source in sources if source["source"] == "odaily"), None)
                delay_minutes = None
                if odaily and odaily["published_at"] and event["first_published_at"]:
                    delay_minutes = round((_parse_datetime(odaily["published_at"]) - _parse_datetime(event["first_published_at"])).total_seconds() / 60, 1)
                items.append({
                    **dict(event),
                    "first_sources": _source_labels(first_sources),
                    "first_source_label": "、".join(_source_labels(first_sources)) if first_sources else "数据不足",
                    "odaily_published_at": odaily["published_at"] if odaily else None,
                    "odaily_delay_minutes": delay_minutes,
                    "sources": sources,
                })
        return {"items": items, "page": page, "pages": max(1, (total + page_size - 1) // page_size), "total": total, "page_size": page_size}

    def import_xlsx(self, path: Path, *, start_date: date, end_date: date) -> dict[str, int]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for the one-time XLSX import") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        headers = [str(cell or "").strip() for cell in next(sheet.iter_rows(values_only=True))]
        required = {"ID", "标题", "操作人", "链接", "发布时间", "阅读量", "是否推送", "推送时间"}
        if not required.issubset(headers):
            raise ValueError("XLSX headers do not match the confirmed Odaily export")
        index = {header: headers.index(header) for header in required}
        facts: list[dict[str, Any]] = []
        for row in sheet.iter_rows(values_only=True):
            published = _parse_datetime(row[index["发布时间"]])
            if published is None or not (start_date <= published.date() < end_date):
                continue
            pushed_at = _parse_datetime(row[index["推送时间"]])
            facts.append({
                "source_item_id": _as_source_item_id(row[index["ID"]]),
                "title": str(row[index["标题"]] or "").strip() or None,
                "operator_raw": str(row[index["操作人"]] or "").strip() or None,
                "source_url": str(row[index["链接"]] or "").strip() or None,
                "published_at": published.isoformat(),
                "view_count": int(row[index["阅读量"]]) if row[index["阅读量"]] is not None else None,
                "is_pushed": 1 if str(row[index["是否推送"]] or "").strip() == "是" else 0,
                "pushed_at": pushed_at.isoformat() if pushed_at else None,
            })
        workbook.close()
        return self.upsert_source_facts(facts, snapshot_at=datetime.fromtimestamp(path.stat().st_mtime, UTC).isoformat())

    def seed_confirmed_week(self, *, actor_email: str = "codex@local") -> None:
        week_start = date(2026, 7, 20)
        schedule = {
            0: ("zoey", "leo", "shark"),
            1: ("harbour", "zoey", "shark"),
            2: ("zoey", "leo", "harbour"),
            3: ("zoey", "harbour", "shark"),
            4: ("zoey", "shark", "harbour"),
            5: ("harbour", "leo"),
            6: ("zoey", "shark"),
        }
        for offset, people in schedule.items():
            duty_date = week_start + timedelta(days=offset)
            mode = "three" if offset < 5 else "two"
            self.save_day_mode({"date": duty_date.isoformat(), "mode": mode}, actor_email=actor_email)
            shift_keys = ("morning", "middle", "late") if mode == "three" else ("morning", "late")
            for shift_key, person_key in zip(shift_keys, people, strict=True):
                self.save_assignment({"date": duty_date.isoformat(), "shift_key": shift_key, "person_key": person_key}, actor_email=actor_email)
        self.save_week_month({"week_start": week_start.isoformat(), "report_month": "2026-07"}, actor_email=actor_email)


def xlsx_fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
